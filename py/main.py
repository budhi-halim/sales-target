import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import io
import copy
from datetime import datetime
from pyscript import ffi, window, document
import re
from collections import defaultdict
from openpyxl.chart import ScatterChart, LineChart, Reference, Series
from openpyxl.chart.marker import Marker
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# Section: Configuration Class
# Holds all configurable settings for columns, formats, and processing rules
class ColumnConfig:
    def __init__(self):
        self.main_column = 'B'  # Main column to check for data presence
        self.extract_columns = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']  # Columns to extract
        self.validation_column = 'K'  # Column to validate for blanks in rows 4-6
        self.key_map = {  # Mapping from column letters to data keys
            'B': 'area',
            'C': 'date',
            'D': 'invoice_no',
            'E': 'customer_name',
            'F': 'product_type',
            'G': 'product_name',
            'H': 'quantity',
            'I': 'unit_price'
        }
        self.output_headers = [  # Headers for the output sales sheet
            'Area',
            'Tanggal',
            'Periode',
            'No. Faktur',
            'Customer',
            'Jenis',
            'Produk',
            'Jumlah',
            'Harga Satuan',
            'Total (IDR)',
            'Kurs',
            'Total (USD)'
        ]
        self.preserve_upper_customer = {  # Words to keep uppercase for customers
            'ABC',
            'BAL',
            'TSG'
        }
        self.preserve_upper_product = {  # Words to keep uppercase for products
            'GMS',
            'HVP',
            'WCI',
            'BBQ',
            'KAN'
        }
        self.area_replacements = {  # Area abbreviations to full names
            'Bdg': 'Bandung',
            'Bgr': 'Bogor',
            'Bks': 'Bekasi',
            'Jkt': 'Jakarta',
            'Lpg': 'Lampung',
            'Mdn': 'Medan',
            'Tgr': 'Tangerang'
        }
        self.idr_format_2 = r'_-"Rp"* #,##0.00_ ;_-"Rp"* -#,##0.00_ ;_-"Rp"* "-"??_ ;_-@_-'
        self.idr_format_0 = r'_-"Rp"* #,##0_ ;_-"Rp"* -#,##0_ ;_-"Rp"* "-"??_ ;_-@_-'
        self.usd_format_2 = r'_-"$"* #,##0.00_ ;_-"$"* -#,##0.00_ ;_-"$"* "-"??_ ;_-@_-'

# Section: Utility Functions
# Helper functions for data processing
def col_to_index(col):
    # Converts column letter to zero-based index
    return ord(col.upper()) - ord('A')

def proper_case(text, preserve_upper=set()):
    if not text:
        return ''
    # Split into alphanum words and non-alphanum separators
    parts = re.findall(r'[a-zA-Z0-9]+|[^a-zA-Z0-9]+', str(text).strip())
    processed_parts = []
    for part in parts:
        if re.match(r'^[a-zA-Z0-9]+$', part):  # word
            # Split into letter and digit subparts
            subparts = re.findall(r'[a-zA-Z]+|[0-9]+', part)
            processed_sub = []
            for sub in subparts:
                if sub.isdigit():
                    processed_sub.append(sub)
                else:  # letters
                    upper_sub = sub.upper()
                    if upper_sub in preserve_upper:
                        processed_sub.append(upper_sub)
                    elif len(sub) <= 2:
                        processed_sub.append(sub)  # preserve case
                    else:
                        processed_sub.append(sub.capitalize())
            processed_parts.append(''.join(processed_sub))
        else:  # separator
            processed_parts.append(part)
    return ''.join(processed_parts)

def process_area(area, replacements):
    # Processes area string with replacements and casing
    if not area:
        return ''
    area = proper_case(area.strip().split()[0])
    return replacements.get(area, area)

# Section: Structure Validation
# Validates input file structure for blanks in validation column
def validate_structure(buffer, filename, config):
    # Checks if specified validation cells are blank
    engine = 'xlrd' if filename.endswith('.xls') else 'openpyxl'
    df = pd.read_excel(buffer, header=None, engine=engine)
    validation_idx = col_to_index(config.validation_column)
    if validation_idx >= df.shape[1]:
        return True  # Column doesn't exist, treat as valid
    cells = [df.iloc[3, validation_idx], df.iloc[4, validation_idx], df.iloc[5, validation_idx]]
    return all(pd.isna(cell) or str(cell).strip() == '' for cell in cells)

# Section: Date Validation
# Validates extracted dates are proper datetime
def validate_dates(data):
    # Checks if all dates in extracted data are valid datetime
    for row in data:
        date_val = row.get('date')
        if date_val and not isinstance(date_val, datetime):
            # Try force convert if str or num
            try:
                if isinstance(date_val, (str, int, float)):
                    pd.to_datetime(date_val)
                else:
                    return False
            except:
                return False
    return True

# Section: Data Extraction
# Extracts relevant data from input file
def extract_data(buffer, filename, config):
    # Reads and extracts data rows based on config, handling date conversion
    engine = 'xlrd' if filename.endswith('.xls') else 'openpyxl'
    df = pd.read_excel(buffer, header=None, engine=engine, skiprows=3)
    main_idx = col_to_index(config.main_column)
    extract_idxs = [col_to_index(col) for col in config.extract_columns]
    data = []
    for _, row in df.iterrows():
        if pd.isna(row[main_idx]) or str(row[main_idx]).strip() == '':
            continue
        row_data = {}
        for i, col in enumerate(config.extract_columns):
            val = row[extract_idxs[i]]
            if pd.isna(val):
                val = ''
            elif col == 'H':
                try:
                    val = float(val)
                except:
                    val = ''
            elif col == 'I':
                try:
                    val = float(val)
                except:
                    val = ''
            elif col == 'C':
                if isinstance(val, (str, int, float)):
                    try:
                        val = pd.to_datetime(val)
                    except:
                        val = val  # Keep original if fail, validate later
            elif col == 'D':
                val = str(val).rstrip('.0') if isinstance(val, float) else str(val)  # Force str, remove .0
            row_data[config.key_map[col]] = val
        data.append(row_data)
    return data

# Section: Data Processing
# Normalizes and processes extracted data
def process_data(combined_data, config):
    # Applies normalization and collects unique periods
    periods = set()
    for row in combined_data:
        row['area'] = process_area(row['area'], config.area_replacements)
        row['customer_name'] = proper_case(row['customer_name'], config.preserve_upper_customer)
        row['product_name'] = proper_case(row['product_name'], config.preserve_upper_product)
        date_value = row.get('date')
        if date_value and isinstance(date_value, datetime):
            periods.add(date_value.strftime('%Y-%m'))
        row['invoice_no'] = str(row['invoice_no']) if row.get('invoice_no') else ''
        row['product_type'] = str(row['product_type']) if row.get('product_type') else ''
        row['quantity'] = row['quantity'] if row.get('quantity') else ''
        row['unit_price'] = row['unit_price'] if row.get('unit_price') else ''
    combined_data = sorted(combined_data, key=lambda x: x.get('date') or datetime.min)
    return sorted(list(periods)), combined_data

# Section: Blank Check
# Checks for blank required fields
def check_blanks(combined_data):
    # Identifies blank cells in required fields
    warnings = []
    key_to_col = {
        'area': 'A',
        'date': 'B',
        'invoice_no': 'D',
        'customer_name': 'E',
        'product_type': 'F',
        'product_name': 'G',
        'quantity': 'H',
        'unit_price': 'I'
    }
    required_keys = list(key_to_col.keys())
    for i, row in enumerate(combined_data, 1):
        for key in required_keys:
            if not row.get(key):
                warnings.append(f"{key_to_col[key]}{i+1}")
    return warnings

# Section: Table Generation
# General function to generate Excel table from 2D data
def generate_table(sheet, headers, data_2d, table_name, col_formats=None, row_formulas=None, start_row=1, start_col=1):
    # Writes headers and data to sheet, applies formats/formulas, creates table
    for rel_c, header in enumerate(headers, 1):
        c = start_col + rel_c - 1
        sheet.cell(start_row, c).value = header
    for r_idx, row_data in enumerate(data_2d, 0):
        r = start_row + r_idx + 1
        for rel_c, val in enumerate(row_data, 1):
            c = start_col + rel_c - 1
            sheet.cell(r, c).value = val
            if col_formats and rel_c in col_formats:
                sheet.cell(r, c).number_format = col_formats[rel_c]
        if row_formulas and row_formulas.get(r_idx):
            for rel_col, formula in row_formulas[r_idx].items():
                c = start_col + rel_col - 1
                sheet.cell(r, c).value = formula
                if col_formats and rel_col in col_formats:
                    sheet.cell(r, c).number_format = col_formats[rel_col]
    last_col = start_col + len(headers) - 1
    last_row = start_row + len(data_2d)
    ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(last_col)}{last_row}"
    tab = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    sheet.add_table(tab)

# New: Sales Target Sheet Generation
# Generates the Sales Target table using grouping and formulas
def generate_sales_target_sheet(wb, grouped_areas_ordered, period_months, config):
    # grouped_areas_ordered: list of groups, each group is list of area names in UI order
    # period_months: list of (year, month) tuples in chronological order
    sheet = wb.create_sheet('Sales Target')

    # Build headers
    headers = ['Area']
    for (y, m) in period_months:
        label = datetime(int(y), int(m), 1).strftime('%b %y')
        headers.append(f'Sales - USD {label}')
    headers.append('Sales - USD Total')
    headers.append('Target - USD Total')
    headers.append('Achieved - USD Total')
    for (y, m) in period_months:
        label = datetime(int(y), int(m), 1).strftime('%b %y')
        headers.append(f'Sales - Qty {label}')
    headers.append('Sales - Qty Total')
    headers.append('Target - Qty Total')
    headers.append('Achieved - Qty Total')

    # Prepare data rows (only groups that have at least one area)
    data_2d = []
    for group in grouped_areas_ordered:
        if not group:
            continue
        label = ', '.join(group)
        # Fill row with blanks; formulas will be applied later
        num_cols = 1 + len(period_months) + 2 + 1 + len(period_months) + 2 + 1  # A + USD months + USD total + target + achieved + Qty months + Qty total + target + achieved
        row = [''] * num_cols
        row[0] = label
        data_2d.append(row)

    # Column formats
    col_formats = {}
    col_formats[1] = '@'
    # USD months + USD total + USD target + USD achieved
    usd_start = 2
    usd_end = usd_start + len(period_months) - 1
    for col in range(usd_start, usd_end + 1):
        col_formats[col] = config.usd_format_2
    usd_total_col = usd_end + 1
    usd_target_col = usd_total_col + 1
    usd_achieved_col = usd_target_col + 1
    col_formats[usd_total_col] = config.usd_format_2
    col_formats[usd_target_col] = config.usd_format_2
    col_formats[usd_achieved_col] = '0%'

    # Qty
    qty_start = usd_achieved_col + 1
    qty_end = qty_start + len(period_months) - 1
    for col in range(qty_start, qty_end + 1):
        col_formats[col] = '#,##0'
    qty_total_col = qty_end + 1
    qty_target_col = qty_total_col + 1
    qty_achieved_col = qty_target_col + 1
    col_formats[qty_total_col] = '#,##0'
    col_formats[qty_target_col] = '#,##0'
    col_formats[qty_achieved_col] = '0%'

    # Row formulas
    row_formulas = {}
    for r_idx, group in enumerate([g for g in grouped_areas_ordered if g]):
        r = r_idx + 2
        # Monthly USD
        for mi, (y, m) in enumerate(period_months):
            c = usd_start + mi
            date_str = f'DATE({int(y)},{int(m)},1)'
            # SUM of SUMIFS across areas in the group
            inner = ','.join([f'SUMIFS(Sales_Data[Total (USD)], Sales_Data[Area], "{area}", Sales_Data[Periode], {date_str})' for area in group])
            formula = f'=SUM({inner})' if inner else '=""'
            row_formulas.setdefault(r_idx, {})[c] = formula
        # USD total
        sum_range = f'{get_column_letter(usd_start)}{r}:{get_column_letter(usd_end)}{r}' if len(period_months) > 0 else ''
        row_formulas.setdefault(r_idx, {})[usd_total_col] = f'=SUM({sum_range})' if sum_range else '=""'
        # USD target left blank (we will leave empty; formatting applied). If formatting of empty not retained, user can input. No default value here.
        # Achieved USD
        row_formulas.setdefault(r_idx, {})[usd_achieved_col] = f'=IFERROR({get_column_letter(usd_total_col)}{r}/{get_column_letter(usd_target_col)}{r},0)'

        # Monthly Qty
        for mi, (y, m) in enumerate(period_months):
            c = qty_start + mi
            date_str = f'DATE({int(y)},{int(m)},1)'
            inner = ','.join([f'SUMIFS(Sales_Data[Jumlah], Sales_Data[Area], "{area}", Sales_Data[Periode], {date_str})' for area in group])
            formula = f'=SUM({inner})' if inner else '=""'
            row_formulas.setdefault(r_idx, {})[c] = formula
        # Qty total
        sum_range_qty = f'{get_column_letter(qty_start)}{r}:{get_column_letter(qty_end)}{r}' if len(period_months) > 0 else ''
        row_formulas.setdefault(r_idx, {})[qty_total_col] = f'=SUM({sum_range_qty})' if sum_range_qty else '=""'
        # Achieved Qty
        row_formulas.setdefault(r_idx, {})[qty_achieved_col] = f'=IFERROR({get_column_letter(qty_total_col)}{r}/{get_column_letter(qty_target_col)}{r},0)'

    generate_table(sheet, headers, data_2d, 'Sales_Target', col_formats, row_formulas)

# Helper: Build period months list from sorted_periods (YYYY-MM)
def build_period_months(sorted_periods):
    months = []
    for p in sorted(sorted_periods):
        y = int(p[:4])
        m = int(p[5:])
        months.append((y, m))
    return months

# Section: Main Processing Functions
# We split into prepare (no workbook) and finalize (creates workbook)

def prepare_files(file_datas):
    message_div = document.getElementById('message')
    message_div.innerHTML = ''

    file_datas = file_datas.to_py()
    if not file_datas:
        message = 'No files uploaded.'
        message_div.innerHTML = message
        return {'message': message, 'type': 'error'}

    config = ColumnConfig()
    invalid_files = []
    invalid_date_files = []
    combined_data = []

    for f in file_datas:
        name = f['name']
        buffer = io.BytesIO(bytes(f['data']))
        if not validate_structure(buffer, name, config):
            invalid_files.append(name)
            continue
        buffer.seek(0)
        data = extract_data(buffer, name, config)
        if not validate_dates(data):
            invalid_date_files.append(name)
            continue
        combined_data.extend(data)

    msg = ''
    if invalid_files:
        msg += '<p>These files do not follow the required format:</p><ul>' + ''.join(f'<li>{f}</li>' for f in invalid_files) + '</ul>'
    if invalid_date_files:
        if msg:
            msg += ''
        msg += '<p>These files have invalid date formatting:</p><ul>' + ''.join(f'<li>{f}</li>' for f in invalid_date_files) + '</ul>'
    if msg:
        message_div.innerHTML = msg
        return {'message': msg, 'type': 'error'}

    sorted_periods, combined_data = process_data(combined_data, config)
    blank_cells = check_blanks(combined_data)

    # Collect unique areas (bench)
    areas = []
    seen = set()
    for row in combined_data:
        a = row.get('area')
        if a and a not in seen:
            seen.add(a)
            areas.append(a)

    years = sorted(set(p[:4] for p in sorted_periods))
    multi_year = len(years) > 1

    # Return minimal info to build UI
    return {
        'message': 'Ready for confirmation',
        'type': 'success',
        'areas': areas,
        'periods': sorted_periods,
        'years': years,
        'multi_year': multi_year,
        'blank_cells': blank_cells,
    }


def finalize_files(file_datas, grouping, proceed_multi_year):
    # Statelessly rebuild combined_data and then create workbook
    prep = prepare_files(file_datas)
    if prep.get('type') == 'error':
        return prep
    if prep.get('multi_year') and not proceed_multi_year:
        return {'message': 'Multiple years detected. Processing cancelled by user.', 'type': 'error'}

    config = ColumnConfig()

    # Rebuild combined_data again
    file_datas = file_datas.to_py()
    combined_data = []
    for f in file_datas:
        name = f['name']
        buffer = io.BytesIO(bytes(f['data']))
        if not validate_structure(buffer, name, config):
            continue
        buffer.seek(0)
        data = extract_data(buffer, name, config)
        if not validate_dates(data):
            continue
        combined_data.extend(data)
    sorted_periods, combined_data = process_data(combined_data, config)

    wb = Workbook()

    # Sheet 1: Sales Data
    sheet1 = wb.active
    sheet1.title = 'Sales Data'
    sales_data_2d = []
    sales_row_formulas = {}
    sales_col_formats = {
        1: '@',
        2: 'dd/mm/yyyy',
        3: 'dd/mm/yyyy',
        4: '@',
        5: '@',
        6: '@',
        7: '@',
        8: '#,##0',
        9: config.idr_format_2,
        10: config.idr_format_2,
        11: config.idr_format_0,
        12: config.usd_format_2
    }
    for i, row in enumerate(combined_data):
        date_value = row.get('date') if isinstance(row.get('date'), datetime) else ''
        sales_data_2d.append([row['area'], date_value, '', row['invoice_no'], row['customer_name'], row['product_type'], row['product_name'], row['quantity'], row['unit_price'], '', '', ''])
        sales_row_formulas[i] = {
            3: f'=DATE(YEAR(B{i+2}), MONTH(B{i+2}), 1)',
            10: f'=PRODUCT(H{i+2},I{i+2})',
            11: f'=VLOOKUP(TEXT(C{i+2}, "YYYY-MM"), \'Exchange Rate\'!A:B, 2, FALSE)',
            12: f'=J{i+2}/K{i+2}'
        }
    generate_table(sheet1, config.output_headers, sales_data_2d, "Sales_Data", sales_col_formats, sales_row_formulas)

    # Sheet 2: Exchange Rate
    sheet2 = wb.create_sheet('Exchange Rate')
    exchange_headers = ['Period', 'Rate']
    exchange_data_2d = [[p, None] for p in sorted_periods]
    exchange_col_formats = {1: '@'}
    generate_table(sheet2, exchange_headers, exchange_data_2d, "Exchange_Rate", exchange_col_formats)

    # Sheet 3: Sales Target
    # grouping is a list of groups; each group: list of areas in desired order
    # Ignore any areas not present in grouping
    period_months = build_period_months(sorted_periods)
    groups_filtered = []
    for grp in grouping.to_py():
        # Keep only non-empty groups and areas as provided (order preserved)
        g = [a for a in grp if a]
        if g:
            groups_filtered.append(g)
    generate_sales_target_sheet(wb, groups_filtered, period_months, config)

    # Sheet 4: Cumulative Percentage based on Sales Target
    # Recompute the same column indexing logic used in generate_sales_target_sheet
    n_months = len(period_months)
    usd_start = 2
    usd_end = usd_start + n_months - 1
    usd_total_col = usd_end + 1
    usd_target_col = usd_total_col + 1
    usd_achieved_col = usd_target_col + 1

    qty_start = usd_achieved_col + 1
    qty_end = qty_start + n_months - 1
    qty_total_col = qty_end + 1
    qty_target_col = qty_total_col + 1
    qty_achieved_col = qty_target_col + 1

    # Create sheet for cumulative percentages
    sheet4 = wb.create_sheet('Cumulative Percentage')

    # Build headers: Area, then Sales - USD MMM YY (for each period), then Sales - Qty MMM YY (for each period)
    headers_cp = ['Area']
    for (y, m) in period_months:
        label = datetime(int(y), int(m), 1).strftime('%b %y')
        headers_cp.append(f'Sales - USD {label}')
    for (y, m) in period_months:
        label = datetime(int(y), int(m), 1).strftime('%b %y')
        headers_cp.append(f'Sales - Qty {label}')

    # Prepare empty data rows (formulas will be applied per row)
    data_2d_cp = []
    for group in groups_filtered:
        # Table uses group label same as Sales Target (A = ', '.join(group))
        data_2d_cp.append([', '.join(group)] + [''] * (2 * n_months))

    # Build row_formulas for each data row: cumulative sums divided by target columns in Sales Target
    row_formulas_cp = {}
    for r_idx, group in enumerate(groups_filtered):
        # Row index in sheet4 (1-based header row at 1): data rows start at row 2
        row_spreadsheet = r_idx + 2

        # Corresponding row in Sales Target is the same row index (Sales Target table also starts at row 1 with header)
        sales_row = row_spreadsheet

        # USD cumulative formulas
        for mi in range(n_months):
            # cumulative columns in Sales Target: usd_start .. (usd_start + mi)
            left_col_letter = get_column_letter(usd_start)
            right_col_letter = get_column_letter(usd_start + mi)
            target_col_letter = get_column_letter(usd_target_col)
            # Build Excel SUM range for the Sales Target row
            sum_range = f"'Sales Target'!{left_col_letter}{sales_row}:{right_col_letter}{sales_row}"
            # Formula: =IFERROR(SUM(range)/'Sales Target'!TargetCell,0)
            formula = f"=IFERROR(SUM({sum_range})/'Sales Target'!{target_col_letter}{sales_row},0)"
            # Place this formula in sheet4 at relative column (2 + mi)
            row_formulas_cp.setdefault(r_idx, {})[1 + mi + 1] = formula  # +1 because headers start at col1 (Area)

        # Qty cumulative formulas (placed after USD columns)
        for mi in range(n_months):
            left_q_col = qty_start
            right_q_col = qty_start + mi
            left_q_letter = get_column_letter(left_q_col)
            right_q_letter = get_column_letter(right_q_col)
            qty_target_letter = get_column_letter(qty_target_col)
            sum_range_q = f"'Sales Target'!{left_q_letter}{sales_row}:{right_q_letter}{sales_row}"
            formula_q = f"=IFERROR(SUM({sum_range_q})/'Sales Target'!{qty_target_letter}{sales_row},0)"
            # relative column index for qty = 1 (area) + n_months + (mi+1)
            rel_col = 1 + n_months + (mi + 1)
            row_formulas_cp.setdefault(r_idx, {})[rel_col] = formula_q

    # Column formats: USD and Qty cumulative as percentage with 0 decimals
    col_formats_cp = {}
    # Area as text
    col_formats_cp[1] = '@'
    # USD cumulative columns (relative columns 2..1+n_months)
    for rel in range(2, 2 + n_months):
        col_formats_cp[rel] = '0%'
    # Qty cumulative columns
    for rel in range(2 + n_months, 2 + 2 * n_months):
        col_formats_cp[rel] = '0%'

    # Create the table
    generate_table(sheet4, headers_cp, data_2d_cp, 'Cumulative_Percentage', col_formats_cp, row_formulas_cp)

    # Very hide sheet4
    sheet4.sheet_state = 'veryHidden'

    sheet3 = wb['Sales Target']
    sheet4 = wb['Cumulative Percentage']  # Reverted to original name with space

    num_data_rows = len([g for g in groups_filtered if g])
    chart_row = 1 + num_data_rows + 2  # header + data + one blank

    if n_months > 0 and num_data_rows > 0:
        for r in range(1, min(6, sheet4.max_row + 1)):
            row_vals = [sheet4.cell(row=r, column=c).value for c in range(1, min(sheet4.max_column + 1, 10))]

        try:
            # --- USD Chart ---
            usd_chart = ScatterChart()
            usd_chart.layout = Layout(
                manualLayout=ManualLayout(
                    x=0,
                    y=0,
                    h=0.8,
                    w=0.5
                )
            )
            usd_chart.auto_title_deleted = False
            usd_chart.scatterStyle = 'lineMarker'
            usd_chart.title = "Sales (USD)"
            usd_chart.x_axis.majorGridlines = None
            usd_chart.y_axis.majorGridlines = None
            usd_chart.x_axis.title = 'Month'
            usd_chart.y_axis.title = 'Cumulative %'
            usd_chart.x_axis.delete = False
            usd_chart.y_axis.delete = False

            # Categories: Month labels from headers (row 1, columns 2 to 1+n_months)
            usd_cat = Reference(sheet4, min_col=2, min_row=1, max_col=1 + n_months, max_row=1)

            usd_series_count = 0
            for r_idx in range(num_data_rows):
                data_row = 2 + r_idx
                title_cell = sheet4.cell(row=data_row, column=1)
                title_val = title_cell.value
                if not title_val:
                    continue

                # Data range: same row, columns 2 to 1+n_months
                yref = Reference(
                    sheet4,
                    min_col=2,
                    min_row=data_row,
                    max_col=1 + n_months,
                    max_row=data_row
                )

                # Only add series if title exists
                if title_val:
                    series = Series(yref, xvalues=usd_cat, title=title_val)
                    series.marker.symbol = 'circle'
                    series.marker.size = 3
                    series.graphicalProperties.line.width = 12500
                    usd_chart.series.append(series)
                    usd_series_count += 1

            usd_chart.width = 20
            usd_chart.height = 10
            anchor = f"A{chart_row}"
            sheet3.add_chart(usd_chart, anchor)

        except Exception as e:
            import traceback
            traceback.print_exc()

        try:
            # --- Qty Chart ---
            qty_chart = ScatterChart()
            qty_chart.layout = Layout(
                manualLayout=ManualLayout(
                    x=0,
                    y=0,
                    h=0.8,
                    w=0.5
                )
            )
            qty_chart.auto_title_deleted = False
            qty_chart.scatterStyle = 'lineMarker'
            qty_chart.title = "Sales (Qty)"
            qty_chart.x_axis.majorGridlines = None
            qty_chart.y_axis.majorGridlines = None
            qty_chart.x_axis.title = 'Month'
            qty_chart.y_axis.title = 'Cumulative %'
            qty_chart.x_axis.delete = False
            qty_chart.y_axis.delete = False

            qty_cat = Reference(sheet4, min_col=2 + n_months, min_row=1, max_col=1 + 2 * n_months, max_row=1)

            qty_series_count = 0
            for r_idx in range(num_data_rows):
                data_row = 2 + r_idx
                title_val_q = sheet4.cell(row=data_row, column=1).value
                if not title_val_q:
                    continue

                yref_q = Reference(
                    sheet4,
                    min_col=2 + n_months,
                    min_row=data_row,
                    max_col=1 + 2 * n_months,
                    max_row=data_row
                )

                if title_val_q:
                    series_q = Series(yref_q, xvalues=qty_cat, title=title_val_q)
                    series_q.marker.symbol = 'circle'
                    series_q.marker.size = 3
                    series_q.graphicalProperties.line.width = 12500
                    qty_chart.series.append(series_q)
                    qty_series_count += 1

            qty_chart.width = 20
            qty_chart.height = 10
            anchor_qty = f"L{chart_row}"
            sheet3.add_chart(qty_chart, anchor_qty)

        except Exception as e:
            import traceback
            traceback.print_exc()

    # Save
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    buffer = list(output.getvalue())

    blank_cells = prep.get('blank_cells', [])
    if blank_cells:
        msg = '<p>Processing complete. Warning: Please check these empty cells in the output:</p><ul>' + ''.join(f'<li>{c}</li>' for c in blank_cells) + '</ul>'
        message_type = 'warning'
    else:
        msg = 'Processing complete. All data processed successfully.'
        message_type = 'success'

    return {'buffer': buffer, 'message': msg, 'type': message_type}

# Bind to window for JS
window.prepare_files = ffi.create_proxy(prepare_files)
window.finalize_files = ffi.create_proxy(finalize_files)